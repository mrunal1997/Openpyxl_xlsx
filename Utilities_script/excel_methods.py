import openpyxl

sheet_path = "TestData/sample.xlsx"
worksheet = openpyxl.load_workbook(sheet_path)
sheet_names = worksheet.sheetnames
# print(sheet_names)
# sheet = worksheet["Sheet"]
# List = list(sheet.values)#[1:]
# print(List)


class BaseData:
    def __init__(self, sheet_name):
        self.sheet = worksheet[sheet_name]


class ExcelMethods(BaseData):

    def get_parametrize_list(self):
        self.del_existing_result_column()
        parametrize_list = list(self.sheet.values)[1:]
        print(parametrize_list)
        return parametrize_list

    def update_result_in_excel(self, t_c_no, status):
        row_number = 0
        self.create_result_column()
        column_number = self.get_result_column_number()
        for col in self.sheet.iter_cols(values_only=True, max_col=1):
            for cell in col:
                row_number += 1
                if cell == t_c_no:
                    break
        if status == True:
            status = 'Pass'
        else:
            status = 'Fail'
        self.sheet.cell(row_number, column_number).value = status
        worksheet.save(sheet_path)

    def del_existing_result_column(self):
        list_with_values = []
        for cell in self.sheet[1]:
            list_with_values.append(cell.value)
        if 'Test Result' in list_with_values:
            self.sheet.delete_cols(self.sheet.max_column)
        worksheet.save(sheet_path)

    def get_result_column_number(self):
        list_with_values = []
        for cell in self.sheet[1]:
            list_with_values.append(cell.value)
        column_number = list_with_values.index('Test Result') + 1
        return column_number

    def create_result_column(self):
        list_with_values = []
        for cell in self.sheet[1]:
            list_with_values.append(cell.value)
        if 'Test Result' in list_with_values:
            pass
        else:
            self.sheet.cell(1, self.sheet.max_column + 1).value = 'Test Result'
