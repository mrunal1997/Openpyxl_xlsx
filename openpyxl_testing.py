from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
sheet.title = "Demo XLSX"

driver = webdriver.Chrome()
driver.get("https://www.rahulshettyacademy.com/AutomationPractice/")

# number_of_rows = driver.find_elements("xpath", "//div[@class='tableFixHead']//tbody/tr")
# number_of_columns = driver.find_elements("xpath", "//div[@class='tableFixHead']//tbody/tr[1]/td")
#
# for i, row in enumerate(number_of_rows, start=1):
#     for j, column in enumerate(number_of_columns, start=1):
#         sheet.cell(i, j).value = driver.find_element("xpath", f"//div[@class='tableFixHead']//tbody/tr[{i}]/td[{j}]").text
#
# wb.save("demo1.xlsx")

number_of_rows = driver.find_elements("xpath", "//table[@class='table-display']//tbody/tr")
number_of_columns = driver.find_elements("xpath", "//table[@class='table-display']//tbody/tr[2]/td")

for i, row in enumerate(number_of_rows, start=1):
    for j, column in enumerate(number_of_columns, start=1):
        sheet.cell(i, j).value = driver.find_element("xpath", f"//table[@id='product']//tbody/tr[{i}]/td[{j}]").text

wb.save("demo1.xlsx")